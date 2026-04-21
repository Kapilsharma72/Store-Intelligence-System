"""
Export endpoints for the Enhanced Web Dashboard.

Provides:
  - POST /api/v1/videos/{video_id}/export/pdf   — PDF report with cover page, summary stats, and data tables
  - POST /api/v1/videos/{video_id}/export/excel — Excel workbook with Summary, Metrics, Journey, Timeseries sheets

Requirements: 16.1, 16.2, 16.3, 16.4, 16.5
"""

from __future__ import annotations

import io
import statistics
from collections import defaultdict
from datetime import datetime, timezone
from typing import Annotated, Any, Dict, List, Optional

import structlog
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.auth import UserContext, get_current_user
from app.database import get_db
from app.models import Event as EventModel
from app.videos import _check_ownership, _get_video_or_404

logger = structlog.get_logger()

router = APIRouter(prefix="/api/v1/videos", tags=["export"])

# ---------------------------------------------------------------------------
# MIME types
# ---------------------------------------------------------------------------
PDF_MIME = "application/pdf"
EXCEL_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ---------------------------------------------------------------------------
# Shared data helpers
# ---------------------------------------------------------------------------

def _ts_aware(ts: datetime) -> datetime:
    if ts.tzinfo is None:
        return ts.replace(tzinfo=timezone.utc)
    return ts


def _get_events(video_id: str, db: Session) -> list:
    return (
        db.query(EventModel)
        .filter(EventModel.metadata_["video_id"].as_string() == video_id)
        .order_by(EventModel.timestamp)
        .all()
    )


def _compute_summary(video_id: str, db: Session) -> Dict[str, Any]:
    """Compute summary statistics for a video."""
    events = _get_events(video_id, db)

    unique_visitors = len({
        ev.visitor_id for ev in events
        if ev.event_type == "ENTRY" and not ev.is_staff
    })
    queue_joins = {
        ev.visitor_id for ev in events
        if ev.event_type == "BILLING_QUEUE_JOIN" and not ev.is_staff
    }
    queue_abandons = sum(
        1 for ev in events
        if ev.event_type == "BILLING_QUEUE_ABANDON" and not ev.is_staff
    )
    queue_depth = max(0, len(queue_joins) - queue_abandons)
    conversion_rate = round(min(1.0, len(queue_joins) / unique_visitors), 4) if unique_visitors > 0 else 0.0

    dwell_values = [
        ev.dwell_ms / 1000.0
        for ev in events
        if ev.event_type == "ZONE_DWELL" and ev.dwell_ms is not None and not ev.is_staff
    ]
    avg_dwell = round(statistics.mean(dwell_values), 2) if dwell_values else 0.0

    staff_count = len({ev.visitor_id for ev in events if ev.is_staff})
    customer_count = len({ev.visitor_id for ev in events if not ev.is_staff})

    return {
        "unique_visitors": unique_visitors,
        "conversion_rate": conversion_rate,
        "avg_dwell_seconds": avg_dwell,
        "queue_depth": queue_depth,
        "abandonment_count": queue_abandons,
        "staff_count": staff_count,
        "customer_count": customer_count,
        "total_events": len(events),
    }


def _compute_journey_rows(video_id: str, db: Session) -> List[Dict[str, Any]]:
    """Return flattened journey rows for export."""
    events = (
        db.query(EventModel)
        .filter(
            EventModel.metadata_["video_id"].as_string() == video_id,
            EventModel.event_type.in_(["ZONE_ENTER", "ZONE_EXIT", "ZONE_DWELL"]),
        )
        .order_by(EventModel.timestamp)
        .all()
    )

    visitor_events: Dict[str, list] = defaultdict(list)
    for ev in events:
        visitor_events[ev.visitor_id].append(ev)

    rows: List[Dict[str, Any]] = []
    for visitor_id, evs in visitor_events.items():
        open_entries: Dict[str, Any] = {}
        for ev in evs:
            if ev.event_type == "ZONE_ENTER":
                open_entries[ev.zone_id] = ev
            elif ev.event_type == "ZONE_EXIT":
                entry_ev = open_entries.pop(ev.zone_id, None)
                entry_ts = entry_ev.timestamp if entry_ev else None
                dwell = None
                if entry_ts is not None:
                    dwell = round((_ts_aware(ev.timestamp) - _ts_aware(entry_ts)).total_seconds(), 2)
                rows.append({
                    "visitor_id": visitor_id,
                    "zone_id": ev.zone_id or "",
                    "entry_timestamp": str(entry_ts) if entry_ts else "",
                    "exit_timestamp": str(ev.timestamp),
                    "dwell_time_seconds": dwell,
                })
            elif ev.event_type == "ZONE_DWELL" and ev.zone_id not in open_entries:
                rows.append({
                    "visitor_id": visitor_id,
                    "zone_id": ev.zone_id or "",
                    "entry_timestamp": str(ev.timestamp),
                    "exit_timestamp": "",
                    "dwell_time_seconds": round(ev.dwell_ms / 1000.0, 2) if ev.dwell_ms else None,
                })
    return rows


def _compute_timeseries_rows(video_id: str, db: Session) -> List[Dict[str, Any]]:
    """Return 30-second interval rows for export."""
    from datetime import timedelta

    events = _get_events(video_id, db)
    if not events:
        return []

    min_ts = _ts_aware(events[0].timestamp)
    max_ts = _ts_aware(events[-1].timestamp)
    interval_secs = 30
    rows: List[Dict[str, Any]] = []

    current_start = min_ts
    while current_start <= max_ts:
        current_end = current_start + timedelta(seconds=interval_secs)

        cumulative_visitors = len({
            ev.visitor_id for ev in events
            if _ts_aware(ev.timestamp) < current_end
            and ev.event_type == "ENTRY" and not ev.is_staff
        })
        cumulative_joins = len({
            ev.visitor_id for ev in events
            if _ts_aware(ev.timestamp) < current_end
            and ev.event_type == "BILLING_QUEUE_JOIN" and not ev.is_staff
        })
        conversion_rate = round(min(1.0, cumulative_joins / cumulative_visitors), 4) if cumulative_visitors > 0 else 0.0

        interval_events = [
            ev for ev in events
            if _ts_aware(ev.timestamp) >= current_start and _ts_aware(ev.timestamp) < current_end
        ]
        joins = sum(1 for ev in interval_events if ev.event_type == "BILLING_QUEUE_JOIN" and not ev.is_staff)
        abandons = sum(1 for ev in interval_events if ev.event_type == "BILLING_QUEUE_ABANDON" and not ev.is_staff)
        queue_depth = max(0, joins - abandons)

        dwell_values = [
            ev.dwell_ms / 1000.0
            for ev in interval_events
            if ev.event_type == "ZONE_DWELL" and ev.dwell_ms is not None and not ev.is_staff
        ]
        avg_dwell = round(statistics.mean(dwell_values), 2) if dwell_values else 0.0

        rows.append({
            "interval_start": str(current_start),
            "interval_end": str(current_end),
            "unique_visitors_cumulative": cumulative_visitors,
            "conversion_rate": conversion_rate,
            "queue_depth": queue_depth,
            "avg_dwell_seconds": avg_dwell,
        })
        current_start = current_end

    return rows


# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

def _generate_pdf(video_id: str, video: Any, summary: Dict[str, Any]) -> bytes:
    """Generate a PDF report using reportlab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontSize=22,
        spaceAfter=12,
        textColor=colors.HexColor("#1a1a2e"),
    )
    heading_style = ParagraphStyle(
        "Heading",
        parent=styles["Heading2"],
        fontSize=14,
        spaceBefore=16,
        spaceAfter=8,
        textColor=colors.HexColor("#16213e"),
    )
    body_style = styles["BodyText"]

    generation_ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    story = []

    # ---- Cover page ----
    story.append(Paragraph("Store Intelligence Report", title_style))
    story.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#0f3460")))
    story.append(Spacer(1, 0.4 * cm))

    cover_data = [
        ["Video ID", video_id],
        ["Filename", video.filename],
        ["Duration (s)", str(video.duration_seconds or "N/A")],
        ["Resolution", video.resolution or "N/A"],
        ["Codec", video.codec or "N/A"],
        ["Store Config", video.store_config or "N/A"],
        ["Upload Time", str(video.upload_timestamp)],
        ["Report Generated", generation_ts],
    ]
    cover_table = Table(cover_data, colWidths=[5 * cm, 11 * cm])
    cover_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#e8f4f8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#f7fbfd")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(cover_table)
    story.append(Spacer(1, 0.8 * cm))

    # ---- Summary Statistics ----
    story.append(Paragraph("Summary Statistics", heading_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#cccccc")))
    story.append(Spacer(1, 0.3 * cm))

    stats_data = [
        ["Metric", "Value"],
        ["Unique Visitors", str(summary["unique_visitors"])],
        ["Conversion Rate", f"{summary['conversion_rate'] * 100:.1f}%"],
        ["Avg Dwell Time (s)", str(summary["avg_dwell_seconds"])],
        ["Queue Depth", str(summary["queue_depth"])],
        ["Abandonment Count", str(summary["abandonment_count"])],
        ["Staff Count", str(summary["staff_count"])],
        ["Customer Count", str(summary["customer_count"])],
        ["Total Events", str(summary["total_events"])],
    ]
    stats_table = Table(stats_data, colWidths=[8 * cm, 8 * cm])
    stats_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3460")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (1, 0), (-1, -1), [colors.white, colors.HexColor("#f0f8ff")]),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(stats_table)
    story.append(Spacer(1, 0.8 * cm))

    # ---- Chart Data Note ----
    story.append(Paragraph("Analytics Data", heading_style))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#cccccc")))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(
        "Detailed analytics data including visitor journeys, time-series metrics, "
        "zone performance, and queue analysis are available in the Excel export. "
        "Interactive charts are available in the web dashboard.",
        body_style,
    ))
    story.append(Spacer(1, 0.4 * cm))

    # Key metrics highlight table
    highlight_data = [
        ["Metric", "Value", "Benchmark"],
        ["Conversion Rate", f"{summary['conversion_rate'] * 100:.1f}%", "Industry avg: 20-30%"],
        ["Avg Dwell Time", f"{summary['avg_dwell_seconds']:.1f}s", "Target: > 60s"],
        ["Queue Depth", str(summary["queue_depth"]), "Target: < 5"],
    ]
    highlight_table = Table(highlight_data, colWidths=[5.5 * cm, 4 * cm, 6.5 * cm])
    highlight_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#16213e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (1, 0), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(highlight_table)

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _generate_excel(
    video_id: str,
    video: Any,
    summary: Dict[str, Any],
    journey_rows: List[Dict[str, Any]],
    timeseries_rows: List[Dict[str, Any]],
) -> bytes:
    """Generate an Excel workbook using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Font,
        PatternFill,
        numbers,
    )
    from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ---- Colour palette ----
    HEADER_FILL = PatternFill("solid", fgColor="0F3460")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBHEADER_FILL = PatternFill("solid", fgColor="16213E")
    SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL = PatternFill("solid", fgColor="F0F8FF")
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

    def _style_header_row(ws, row: int, col_count: int) -> None:
        for c in range(1, col_count + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    def _auto_width(ws) -> None:
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # ================================================================
    # Sheet 1: Summary
    # ================================================================
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary.append(["Store Intelligence Report — Summary"])
    ws_summary["A1"].font = Font(bold=True, size=16, color="1A1A2E")
    ws_summary.merge_cells("A1:C1")
    ws_summary.append([])

    ws_summary.append(["Video Metadata"])
    ws_summary["A3"].font = Font(bold=True, size=12, color="0F3460")
    ws_summary.merge_cells("A3:C3")

    meta_rows = [
        ("Video ID", video_id),
        ("Filename", video.filename),
        ("Duration (s)", video.duration_seconds or "N/A"),
        ("Resolution", video.resolution or "N/A"),
        ("Codec", video.codec or "N/A"),
        ("Store Config", video.store_config or "N/A"),
        ("Upload Time", str(video.upload_timestamp)),
        ("Report Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]
    for label, value in meta_rows:
        ws_summary.append([label, value])

    ws_summary.append([])
    ws_summary.append(["Key Metrics"])
    metrics_header_row = ws_summary.max_row
    ws_summary[f"A{metrics_header_row}"].font = Font(bold=True, size=12, color="0F3460")
    ws_summary.merge_cells(f"A{metrics_header_row}:C{metrics_header_row}")

    ws_summary.append(["Metric", "Value", "Notes"])
    header_row = ws_summary.max_row
    _style_header_row(ws_summary, header_row, 3)

    metrics_data = [
        ("Unique Visitors", summary["unique_visitors"], ""),
        ("Conversion Rate", summary["conversion_rate"], "Ratio 0-1"),
        ("Avg Dwell Time (s)", summary["avg_dwell_seconds"], ""),
        ("Queue Depth", summary["queue_depth"], "Net joins minus abandons"),
        ("Abandonment Count", summary["abandonment_count"], ""),
        ("Staff Count", summary["staff_count"], ""),
        ("Customer Count", summary["customer_count"], ""),
        ("Total Events", summary["total_events"], ""),
    ]
    for i, (metric, value, note) in enumerate(metrics_data):
        ws_summary.append([metric, value, note])
        row = ws_summary.max_row
        if i % 2 == 1:
            for c in range(1, 4):
                ws_summary.cell(row=row, column=c).fill = ALT_FILL

    # Conditional formatting: conversion rate column (B) — color scale
    cr_start = header_row + 1
    cr_end = header_row + 1  # only conversion rate row
    ws_summary.conditional_formatting.add(
        f"B{cr_start}:B{cr_end}",
        ColorScaleRule(
            start_type="num", start_value=0, start_color="FF0000",
            end_type="num", end_value=1, end_color="00FF00",
        ),
    )

    _auto_width(ws_summary)

    # ================================================================
    # Sheet 2: Metrics (event-level aggregates by event type)
    # ================================================================
    ws_metrics = wb.create_sheet("Metrics")
    ws_metrics.append(["Event Type", "Count", "Staff", "Customer"])
    _style_header_row(ws_metrics, 1, 4)

    # Use the summary data we have; for detailed breakdown we note it's in the DB
    # Add a note row
    ws_metrics.append(["Note: Detailed per-event-type breakdown requires DB query at export time.", "", "", ""])
    ws_metrics.append([])
    ws_metrics.append(["Metric", "Value"])
    _style_header_row(ws_metrics, ws_metrics.max_row, 2)

    for metric, value, _ in metrics_data:
        ws_metrics.append([metric, value])

    # Conditional formatting on Value column (B) for numeric rows
    data_start = 4
    data_end = ws_metrics.max_row
    ws_metrics.conditional_formatting.add(
        f"B{data_start}:B{data_end}",
        DataBarRule(start_type="min", start_value=0, end_type="max", end_value=None,
                    color="638EC6"),
    )

    _auto_width(ws_metrics)

    # ================================================================
    # Sheet 3: Journey
    # ================================================================
    ws_journey = wb.create_sheet("Journey")
    journey_headers = ["Visitor ID", "Zone ID", "Entry Timestamp", "Exit Timestamp", "Dwell Time (s)"]
    ws_journey.append(journey_headers)
    _style_header_row(ws_journey, 1, len(journey_headers))

    for i, row in enumerate(journey_rows):
        ws_journey.append([
            row["visitor_id"],
            row["zone_id"],
            row["entry_timestamp"],
            row["exit_timestamp"],
            row["dwell_time_seconds"],
        ])
        if i % 2 == 1:
            for c in range(1, len(journey_headers) + 1):
                ws_journey.cell(row=i + 2, column=c).fill = ALT_FILL

    # Conditional formatting on dwell time column (E)
    if journey_rows:
        ws_journey.conditional_formatting.add(
            f"E2:E{len(journey_rows) + 1}",
            ColorScaleRule(
                start_type="min", start_value=None, start_color="FFFFFF",
                end_type="max", end_value=None, end_color="4472C4",
            ),
        )

    _auto_width(ws_journey)

    # ================================================================
    # Sheet 4: Timeseries
    # ================================================================
    ws_ts = wb.create_sheet("Timeseries")
    ts_headers = [
        "Interval Start", "Interval End",
        "Unique Visitors (Cumulative)", "Conversion Rate",
        "Queue Depth", "Avg Dwell (s)",
    ]
    ws_ts.append(ts_headers)
    _style_header_row(ws_ts, 1, len(ts_headers))

    for i, row in enumerate(timeseries_rows):
        ws_ts.append([
            row["interval_start"],
            row["interval_end"],
            row["unique_visitors_cumulative"],
            row["conversion_rate"],
            row["queue_depth"],
            row["avg_dwell_seconds"],
        ])
        if i % 2 == 1:
            for c in range(1, len(ts_headers) + 1):
                ws_ts.cell(row=i + 2, column=c).fill = ALT_FILL

    # Conditional formatting: conversion rate (D) and queue depth (E)
    if timeseries_rows:
        n = len(timeseries_rows)
        ws_ts.conditional_formatting.add(
            f"D2:D{n + 1}",
            ColorScaleRule(
                start_type="min", start_value=None, start_color="FF0000",
                end_type="max", end_value=None, end_color="00FF00",
            ),
        )
        ws_ts.conditional_formatting.add(
            f"E2:E{n + 1}",
            ColorScaleRule(
                start_type="min", start_value=None, start_color="00FF00",
                end_type="max", end_value=None, end_color="FF0000",
            ),
        )

    _auto_width(ws_ts)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# POST /api/v1/videos/{video_id}/export/pdf
# ---------------------------------------------------------------------------

@router.post("/{video_id}/export/pdf")
async def export_pdf(
    video_id: str,
    current_user: Annotated[UserContext, Depends(get_current_user)] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Generate and return a PDF report for the given video.

    Includes:
    - Cover page with video metadata and generation timestamp
    - Summary statistics section
    - Key metrics highlight table

    Requirements: 16.1, 16.3, 16.5
    """
    video = _get_video_or_404(video_id, db)
    _check_ownership(video, current_user)

    summary = _compute_summary(video_id, db)
    pdf_bytes = _generate_pdf(video_id, video, summary)

    safe_filename = video.filename.replace(" ", "_").replace("/", "_")
    disposition = f'attachment; filename="{safe_filename}_{video_id[:8]}_report.pdf"'

    logger.info("export_pdf", video_id=video_id, exported_by=current_user.username)

    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type=PDF_MIME,
        headers={"Content-Disposition": disposition},
    )


# ---------------------------------------------------------------------------
# POST /api/v1/videos/{video_id}/export/excel
# ---------------------------------------------------------------------------

@router.post("/{video_id}/export/excel")
async def export_excel(
    video_id: str,
    current_user: Annotated[UserContext, Depends(get_current_user)] = None,
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """
    Generate and return an Excel workbook for the given video.

    Sheets: Summary, Metrics, Journey, Timeseries
    Applies conditional formatting on key metric columns.

    Requirements: 16.2, 16.4, 16.5
    """
    video = _get_video_or_404(video_id, db)
    _check_ownership(video, current_user)

    summary = _compute_summary(video_id, db)
    journey_rows = _compute_journey_rows(video_id, db)
    timeseries_rows = _compute_timeseries_rows(video_id, db)

    excel_bytes = _generate_excel(video_id, video, summary, journey_rows, timeseries_rows)

    safe_filename = video.filename.replace(" ", "_").replace("/", "_")
    disposition = f'attachment; filename="{safe_filename}_{video_id[:8]}_report.xlsx"'

    logger.info("export_excel", video_id=video_id, exported_by=current_user.username)

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type=EXCEL_MIME,
        headers={"Content-Disposition": disposition},
    )
