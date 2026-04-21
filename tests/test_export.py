"""
Unit tests for app/export.py — PDF and Excel export endpoints.

Tests cover:
  - POST /api/v1/videos/{video_id}/export/pdf
  - POST /api/v1/videos/{video_id}/export/excel
  - PDF is a valid PDF binary with expected sections
  - Excel workbook has correct sheet names and headers
  - Content-Disposition attachment header is set
  - Correct MIME types are returned
  - 404 for unknown video
  - 403 for unauthorized access

Requirements: 16.1, 16.2, 16.3, 16.4, 16.5
"""

import io
import uuid
from datetime import datetime, timezone

import pytest
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.auth import UserContext, get_current_user
from app.database import Base, get_db
from app.main import app
from app.videos import Video

# ---------------------------------------------------------------------------
# In-memory test database
# ---------------------------------------------------------------------------

TEST_DB_URL = "sqlite:///:memory:"

_engine = create_engine(
    TEST_DB_URL,
    connect_args={"check_same_thread": False},
    poolclass=StaticPool,
)
_TestSession = sessionmaker(autocommit=False, autoflush=False, bind=_engine)


def _override_get_db():
    db = _TestSession()
    try:
        yield db
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def reset_db():
    Base.metadata.drop_all(bind=_engine)
    Base.metadata.create_all(bind=_engine)
    yield


@pytest.fixture
def db():
    session = _TestSession()
    try:
        yield session
    finally:
        session.close()


def _make_client(username: str = "user", role: str = "user") -> TestClient:
    user_ctx = UserContext(username=username, role=role)
    app.dependency_overrides[get_db] = _override_get_db
    app.dependency_overrides[get_current_user] = lambda: user_ctx
    return TestClient(app)


def _insert_video(db, username: str = "user") -> str:
    """Insert a minimal Video row and return its ID."""
    video_id = str(uuid.uuid4())
    video = Video(
        id=video_id,
        filename="test_video.mp4",
        filepath=f"/videos/{video_id}.mp4",
        duration_seconds=120.0,
        resolution="1920x1080",
        codec="h264",
        file_size_bytes=1024 * 1024,
        store_config="store_001",
        status="completed",
        upload_timestamp=datetime.now(timezone.utc),
        uploaded_by=username,
    )
    db.add(video)
    db.commit()
    return video_id


# ---------------------------------------------------------------------------
# PDF export tests
# ---------------------------------------------------------------------------


class TestPDFExport:
    def test_pdf_returns_200(self, db):
        """POST /export/pdf returns HTTP 200 for a valid video."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        assert resp.status_code == 200

    def test_pdf_mime_type(self, db):
        """PDF response has application/pdf content type."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        assert resp.headers["content-type"] == "application/pdf"

    def test_pdf_content_disposition_attachment(self, db):
        """PDF response has Content-Disposition: attachment header."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        disposition = resp.headers.get("content-disposition", "")
        assert "attachment" in disposition
        assert ".pdf" in disposition

    def test_pdf_is_valid_pdf_binary(self, db):
        """Response body starts with the PDF magic bytes %PDF."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        assert resp.content[:4] == b"%PDF", "Response is not a valid PDF"

    def test_pdf_has_nonzero_size(self, db):
        """Generated PDF is not empty."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        assert len(resp.content) > 1024, "PDF is suspiciously small"

    def test_pdf_404_for_unknown_video(self):
        """Returns 404 when video does not exist."""
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{uuid.uuid4()}/export/pdf")
        assert resp.status_code == 404

    def test_pdf_403_for_wrong_user(self, db):
        """Returns 403 when a different user tries to export."""
        video_id = _insert_video(db, username="owner")
        client = _make_client(username="other_user")
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        assert resp.status_code == 403

    def test_pdf_admin_can_export_any_video(self, db):
        """Admin can export any video regardless of ownership."""
        video_id = _insert_video(db, username="owner")
        client = _make_client(username="admin", role="admin")
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        assert resp.status_code == 200

    def test_pdf_filename_in_disposition(self, db):
        """Content-Disposition includes the video filename."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/pdf")
        disposition = resp.headers.get("content-disposition", "")
        # filename should contain part of the video filename or video_id prefix
        assert "test_video" in disposition or video_id[:8] in disposition


# ---------------------------------------------------------------------------
# Excel export tests
# ---------------------------------------------------------------------------


class TestExcelExport:
    def test_excel_returns_200(self, db):
        """POST /export/excel returns HTTP 200 for a valid video."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        assert resp.status_code == 200

    def test_excel_mime_type(self, db):
        """Excel response has correct OOXML MIME type."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        assert "spreadsheetml" in resp.headers["content-type"]

    def test_excel_content_disposition_attachment(self, db):
        """Excel response has Content-Disposition: attachment header."""
        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        disposition = resp.headers.get("content-disposition", "")
        assert "attachment" in disposition
        assert ".xlsx" in disposition

    def test_excel_is_valid_xlsx(self, db):
        """Response body is a valid Excel workbook (ZIP-based OOXML)."""
        from openpyxl import load_workbook

        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb is not None

    def test_excel_sheet_names(self, db):
        """Workbook contains Summary, Metrics, Journey, and Timeseries sheets."""
        from openpyxl import load_workbook

        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Summary" in wb.sheetnames
        assert "Metrics" in wb.sheetnames
        assert "Journey" in wb.sheetnames
        assert "Timeseries" in wb.sheetnames

    def test_excel_summary_sheet_has_video_id(self, db):
        """Summary sheet contains the video ID."""
        from openpyxl import load_workbook

        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Summary"]
        all_values = [str(cell.value or "") for row in ws.iter_rows() for cell in row]
        assert any(video_id in v for v in all_values), "Video ID not found in Summary sheet"

    def test_excel_journey_sheet_headers(self, db):
        """Journey sheet has expected column headers."""
        from openpyxl import load_workbook

        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Journey"]
        headers = [cell.value for cell in ws[1]]
        assert "Visitor ID" in headers
        assert "Zone ID" in headers
        assert "Dwell Time (s)" in headers

    def test_excel_timeseries_sheet_headers(self, db):
        """Timeseries sheet has expected column headers."""
        from openpyxl import load_workbook

        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Timeseries"]
        headers = [cell.value for cell in ws[1]]
        assert "Interval Start" in headers
        assert "Conversion Rate" in headers

    def test_excel_404_for_unknown_video(self):
        """Returns 404 when video does not exist."""
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{uuid.uuid4()}/export/excel")
        assert resp.status_code == 404

    def test_excel_403_for_wrong_user(self, db):
        """Returns 403 when a different user tries to export."""
        video_id = _insert_video(db, username="owner")
        client = _make_client(username="other_user")
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        assert resp.status_code == 403

    def test_excel_admin_can_export_any_video(self, db):
        """Admin can export any video regardless of ownership."""
        video_id = _insert_video(db, username="owner")
        client = _make_client(username="admin", role="admin")
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        assert resp.status_code == 200

    def test_excel_metrics_sheet_has_data(self, db):
        """Metrics sheet contains at least one data row beyond the header."""
        from openpyxl import load_workbook

        video_id = _insert_video(db)
        client = _make_client()
        resp = client.post(f"/api/v1/videos/{video_id}/export/excel")
        wb = load_workbook(io.BytesIO(resp.content))
        ws = wb["Metrics"]
        # Should have more than 1 row (header + data)
        assert ws.max_row > 1
